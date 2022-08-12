from Library.autotestscreen import mainscreen
from Library.cmdaction import *
from Library.times import *
import threading
from Library.testreport import reprot
import openpyxl
from openpyxl.styles import PatternFill
from Library.caplscript import getversion
# from Tools.programswitch import powercontor
# from Tools.powerIO import power_contor


green_fill = PatternFill(fgColor='00b050', fill_type='solid')
red_fill = PatternFill(fgColor='ff0000', fill_type='solid')

class PerformTest(object):

    def __init__(self):
        mainscreen.ScreenButton(button_position={"x": 540, "y": 500},
                                button_param={"width": 8, "height": 2, "text": "结束", "bg": "gray", "command": quit})

        mainscreen.ScreenButton(button_param={"width": 8, "height": 2, "text": "开始",
                                              "bg": "lightskyblue", "command": self.startupdate})
        mainscreen.LabelText(textsize={"x": 40, "y": 300}, labtext="用例路径")
        mainscreen.LabelText(textsize={"x": 40, "y": 520}, labtext="成功数")
        mainscreen.LabelText(textsize={"x": 160, "y": 520}, labtext="成功率")
        mainscreen.LabelText(textsize={"x": 40, "y": 460}, labtext="任务ID")
        mainscreen.LabelText(textsize={"x": 280, "y": 520}, labtext="失败数")
        mainscreen.LabelText(textsize={"x": 400, "y": 520}, labtext="总数")
        mainscreen.LabelText(labtext="图片显示", fontsize=13)
        mainscreen.LabelText(textsize={"x": 700, "y": 30}, labtext="日志打印", fontsize=13)
        mainscreen.LabelText(textsize={"x": 40, "y": 400}, labtext="用例标题")
        mainscreen.MainLoop()

    def mainupdateECU(self):
        s_time = strftime(fmt="%Y-%m-%d %H: %M: %S")
        main_html = ""
        campaignnumber = ''
        resultlist = []
        self.pass_list = []
        self.fail_list = []
        self.c_ecuversion = ""
        self.main_html = ""
        log.info("电源已全部连接，初始化设备中，请勿操作")
        # mainscreen.InsertText(f'{strftime()} 电源已全部连接，初始化设备中，请勿操作')
        # powercontor.contor_on()
        # power_contor("connect", "all")
        # mainscreen.InsertText(f'{strftime()} 打开所有电源开关')
        # sleep(60)
        # adbInit()
        log.info("设备初始化完成")
        mainscreen.InsertText(f'{strftime()} 设备初始化完成')
        self.testcasepath = mainscreen.gettestcasepath()
        sheetname = mainscreen.getsheetname()
        self.excel = openpyxl.load_workbook(self.testcasepath)
        self.sheet = self.excel[sheetname]
        self.i = 1
        for self.value in self.sheet.values:
            self.starttime = strftime(fmt="%Y-%m-%d %H: %M: %S")
            if type(self.value[0]) is int:
                self.i += 1
                data = {}
                data['value1'] = self.value[3]
                data["value2"] = self.value[4]
                for k in list(data.keys()):
                    if data[k] is None:
                        del data[k]
                if self.value[2] == "Web_createcampaign":
                    self.c_ecuversion = getversion(data['value1'])
                    if self.c_ecuversion != None:
                        mainscreen.InsertText(f"{strftime()} {data['value1']}版本获取成功")
                        mainscreen.InsertText(f"{strftime()} {data['value1']}的版本为：{self.c_ecuversion}")
                        # 显示用例标题
                        mainscreen.showcasetitle(f"{self.value[5]}")
                    else:
                        mainscreen.InsertText(f"{strftime()} {data['value1']}版本获取失败：{self.c_ecuversion}")
                        log.error("Version number acquisition failed")
                    # 根据当前版本决定部署的任务
                    if self.value[4] == "不同版本":
                        if self.c_ecuversion == var.V_TBOX_A.split("_")[-2]:
                            self.ecupackge = var.V_TBOX_B
                        else:
                            self.ecupackge = var.V_TBOX_A
                    elif self.value[4] == "相同版本":
                        if self.c_ecuversion == var.V_TBOX_A.split("_")[-2]:
                            self.ecupackge = var.V_TBOX_A
                        else:
                            self.ecupackge = var.V_TBOX_B

                    else:
                        mainscreen.InsertText(f"{strftime()} 测试用例编号: {self.value[0]} 参数输入错误，任务结束")
                        log.error("测试用例编号: {value[0]} 参数输入错误，任务结束")
                        break
                    module = __import__(f"newotatask")
                    mainscreen.InsertText(f'{strftime()} 开始部署{data["value1"]}的升级任务')
                    status = getattr(module, self.value[2])(**data,ecupackage=self.ecupackge)
                    if status:
                        tasktime, campaignnumber = status
                        self.casepass()
                        mainscreen.showID(f"{campaignnumber}")
                        mainscreen.InsertText(f'{strftime()} 升级任务部署成功，任务ID为:{campaignnumber}')
                    else:
                        self.casefail()
                        mainscreen.InsertText(f'{strftime()} 升级任务部署失败')
                        continue
                elif "Car_" in self.value[2]:
                    module = __import__(f"checkcars")
                    status = getattr(module, self.value[2])(**data)
                    # 显示用例标题
                    mainscreen.showcasetitle(f"{self.value[5]}")
                    if status:
                        self.casepass()
                    else:
                        self.casefail()
                elif "Request_" in self.value[2]:
                    module = __import__("requestfunction")
                    status = getattr(module, self.value[2])(**data,campaignid=campaignnumber)
                    # 显示用例标题
                    mainscreen.showcasetitle(f"{self.value[5]}")
                    if status:
                        self.casepass()
                    else:
                        self.casefail()
                elif "Log_" in self.value[2]:
                    module = __import__(f"checklogs")
                    status = getattr(module, self.value[2])(**data, campaignid=campaignnumber)
                    # 显示用例标题
                    mainscreen.showcasetitle(f"{self.value[5]}")
                    if status:
                        self.casepass()
                    else:
                        self.casefail()
                elif "sendpolicy" in self.value[2]:
                    module = __import__(f"caplscript")
                    if data == {}:
                        status = getattr(module, self.value[2])()
                    else:
                        for key in data:
                            excelvalue = str(data[key]).split("=")
                            dictargv = {}
                            dictargv[excelvalue[0]] = excelvalue[1]
                        status = getattr(module, self.value[2])(**dictargv)
                    mainscreen.showcasetitle(f"{self.value[5]}")
                    if status:
                        self.casepass()
                    else:
                        self.casefail()
                else:
                    print(f"tiaoguo: {self.value[2]}")
                casenamber = self.value[0]
                casetitle = self.value[1]
                casedescribe = self.value[5]
                taskid = campaignnumber
                failurestatus = None
                if self.caseresult == "PASS":
                    main_html = main_html + reprot.pass_html(casenamber,taskid,casetitle,casedescribe,self.starttime,
                                                             self.endtime)
                else:
                    main_html = main_html + reprot.fail_html(casenamber, taskid, casetitle, casedescribe, self.starttime,
                                                             self.endtime, failurestatus)
        p_time = strftime(fmt="%Y-%m-%d %H: %M: %S")
        one_cols = [val for val in self.sheet.columns][9]
        for col in one_cols:
            resultlist.append(col.value)
        del resultlist[0]
        passnamber = resultlist.count("PASS")
        failnamber = resultlist.count("FAIL")
        reprot.smoke_test(passnamber,failnamber,main_html,s_time,p_time)
        log.info("打开测试报告")
        mainscreen.InsertText(f'{strftime()} 测试用例全部执行完成，打开测试报告')
        self.excel.close()
        os.system(rf"start {self.testcasepath}")
        mainscreen.InsertText(f'{strftime()} 测试用例全部执行完成，打开测试用例')
        log.info("打开测试用例")
        # power_contor("disconnect", "all")
        # mainscreen.InsertText(f'{strftime()} 关闭所有电源开关')
        # powercontor.contor_off()
        # mainscreen.InsertText(f'{strftime()} 关闭程控电源开关')

    def casepass(self):
        self.caseresult = "PASS"
        self.endtime = strftime(fmt="%Y-%m-%d %H: %M: %S")
        self.sheet.cell(row=self.i, column=8).value = self.starttime
        self.sheet.cell(row=self.i, column=9).value = self.endtime
        self.sheet.cell(row=self.i, column=10).value = "PASS"
        self.sheet.cell(row=self.i, column=10).fill = green_fill
        mainscreen.InsertText(f'{strftime()} 测试用例<<{self.value[5]}>>执行完成，结果为: PASS')
        self.excel.save(self.testcasepath)
        self.pass_list.append(self.value[0])
        self.pass_munber = len(self.pass_list)
        self.fail_munber = len(self.fail_list)
        mainscreen.showsuccess(self.pass_munber)
        mainscreen.showfail(self.fail_munber)
        mainscreen.showrate(self.pass_munber, self.fail_munber)
        mainscreen.showtotal(self.pass_munber, self.fail_munber)

    def casefail(self):
        self.caseresult = "FAIL"
        self.endtime = strftime(fmt="%Y-%m-%d %H: %M: %S")
        self.sheet.cell(row=self.i, column=8).value = self.starttime
        self.sheet.cell(row=self.i, column=9).value = self.endtime
        self.sheet.cell(row=self.i, column=10).value = "FAIL"
        self.sheet.cell(row=self.i, column=10).fill = red_fill
        self.excel.save(self.testcasepath)
        self.fail_list.append(self.value[0])
        self.fail_munber = len(self.fail_list)
        self.pass_munber = len(self.pass_list)
        mainscreen.showsuccess(self.pass_munber)
        mainscreen.showfail(self.fail_munber)
        mainscreen.showrate(self.pass_munber, self.fail_munber)
        mainscreen.showtotal(self.pass_munber, self.fail_munber)
        mainscreen.InsertText(f'{strftime()} 测试用例<<{self.value[5]}>>执行完成，结果为: FAILED')

    def startupdate(self):
        startupgrade = threading.Thread(target=self.mainupdateECU)
        startupgrade.start()


if __name__ == '__main__':
    a = PerformTest()